from decimal import Decimal
import pandas as pd
import openpyxl
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.db import transaction
from django.db.models import Q, Sum, F
from django.http import JsonResponse
from django.forms import inlineformset_factory
from .models import Producto, Categoria, Cliente, Sale, SaleItem, Caja
from .forms import ProductoForm, SaleForm, SaleItemForm
from django.db import transaction
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Producto, Sale, SaleItem
from .forms import SaleForm
from decimal import Decimal, InvalidOperation

# --------------------------
# HISTORIAL DE CAJA
# --------------------------
def historial_caja(request):
    cajas = Caja.objects.all()
    return render(request, "inventario/historial_caja.html", {"cajas": cajas})


# --------------------------
# AUTENTICACIÓN
# --------------------------
def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario registrado correctamente.")
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})
    

def logout_view(request):
    logout(request)
    return redirect('login')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('inicio')

    form = AuthenticationForm(request, data=request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('inicio')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, "inventario/login.html", {"form": form})


# --------------------------
# IMPORTAR EXCEL
# --------------------------
@login_required
def importar_excel(request):
    if request.method == "POST":
        archivo_excel = request.FILES.get("archivo")
        if not archivo_excel:
            messages.error(request, "Por favor, selecciona un archivo Excel.")
            return redirect("importar_excel")

        try:
            # ⚠️ Esto borra todos los productos antes de importar
            # Si quieres que sume, quítalo
            Producto.objects.all().delete()

            workbook = openpyxl.load_workbook(archivo_excel)

            for hoja in workbook.sheetnames:
                categoria_nombre = hoja.strip().upper()
                categoria_obj, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)
                hoja_excel = workbook[hoja]

                for fila in hoja_excel.iter_rows(min_row=2, values_only=True):
                    try:
                        nombre       = str(fila[0]).strip() if fila[0] else ""
                        marca        = str(fila[1]).strip() if fila[1] else ""
                        cantidad_raw = fila[2]
                        unidad_medida= str(fila[3]).strip() if fila[3] else ""
                        precio_raw   = fila[4]

                        if not nombre:  # si no hay nombre, no se guarda
                            continue

                        # --- cantidad ---
                        try:
                            cantidad = int(cantidad_raw) if cantidad_raw not in [None, ""] else 0
                        except (ValueError, TypeError):
                            cantidad = 0

                        # --- precio ---
                        valor = str(precio_raw).strip() if precio_raw is not None else "0"
                        if valor in ["", "nan", None]:
                            precio = Decimal("0")
                        else:
                            try:
                                valor = valor.replace("S/", "").replace(",", "").strip()
                                precio = Decimal(valor)
                            except InvalidOperation:
                                precio = Decimal("0")

                        Producto.objects.create(
                            nombre=nombre,
                            marca=marca,
                            categoria=categoria_obj,
                            cantidad=cantidad,
                            unidad_medida=unidad_medida,
                            precio_venta=precio,
                            precio_compra=0,
                            porcentaje_ganancia=30
                        )

                    except Exception as e:
                        # Si una fila falla, la saltamos pero seguimos importando las demás
                        print(f"Fila con error: {fila} -> {e}")
                        continue

            # Eliminar categorías vacías
            for cat in Categoria.objects.all():
                if not Producto.objects.filter(categoria=cat).exists():
                    cat.delete()

            messages.success(request, "Productos importados correctamente.")
            return redirect("lista_productos")

        except Exception as e:
            messages.error(request, f"Error al importar: {str(e)}")
            return redirect("importar_excel")

    return render(request, "inventario/importar_excel.html")

# --------------------------
# LISTA PRODUCTOS
# --------------------------
@login_required
def lista_productos(request):
    query = request.GET.get('q', '')
    categorias = Categoria.objects.all().order_by("nombre")
    valor_total_inventario = 0
    productos_por_categoria = {}

    for categoria in categorias:
        productos_categoria = Producto.objects.filter(categoria=categoria).order_by("nombre")
        productos_list = []

        for p in productos_categoria:
            precio_compra = p.precio_compra or 0
            precio_venta = p.precio_venta or 0
            cantidad = p.cantidad or 0

            total_inversion = precio_compra * cantidad
            ganancia = (precio_venta - precio_compra) * cantidad
            porcentaje_ganancia = ((precio_venta - precio_compra) / precio_compra * 100) if precio_compra > 0 else 0

            valor_total_inventario += total_inversion

            productos_list.append({
                'obj': p,
                'total_inversion': total_inversion,
                'ganancia': ganancia,
                'porcentaje_ganancia': porcentaje_ganancia,
            })

        productos_por_categoria[categoria] = productos_list

    productos_filtrados = []
    if query:
        qs = Producto.objects.filter(Q(nombre__icontains=query) | Q(marca__icontains=query))
        for p in qs:
            precio_compra = p.precio_compra or 0
            precio_venta = p.precio_venta or 0
            cantidad = p.cantidad or 0

            total_inversion = precio_compra * cantidad
            ganancia = (precio_venta - precio_compra) * cantidad
            porcentaje_ganancia = ((precio_venta - precio_compra) / precio_compra * 100) if precio_compra > 0 else 0

            productos_filtrados.append({
                'obj': p,
                'total_inversion': total_inversion,
                'ganancia': ganancia,
                'porcentaje_ganancia': porcentaje_ganancia,
            })

    context = {
        'categorias': categorias,
        'productos_por_categoria': productos_por_categoria,
        'valor_total_inventario': valor_total_inventario,
        'query': query,
        'productos_filtrados': productos_filtrados,
    }

    return render(request, 'inventario/lista_productos.html', context)


# --------------------------
# CRUD PRODUCTO
# --------------------------
@login_required
def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" agregado correctamente.')
            return redirect('lista_productos')
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        form = ProductoForm()

    return render(request, 'inventario/agregar_producto.html', {'form': form})


@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == "POST":
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'inventario/editar_producto.html', {'form': form})


@login_required
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    nombre_producto = producto.nombre

    if request.method == 'POST':
        producto.delete()
        messages.success(request, f'El producto "{nombre_producto}" fue eliminado correctamente.')
        return redirect('lista_productos')

    return render(request, 'inventario/confirmar_eliminar.html', {'producto': producto})


# --------------------------
# API PRODUCTO
# --------------------------
def producto_api(request, pk):
    try:
        producto = Producto.objects.get(pk=pk)
        data = {
            'id': producto.id,
            'nombre': producto.nombre,
            'precio': producto.precio_venta,
            'stock': producto.cantidad,
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)


# --------------------------
# VENTAS
# --------------------------
SaleItemFormSet = inlineformset_factory(
    Sale, SaleItem,
    form=SaleItemForm,
    extra=1,
    can_delete=True
)

@transaction.atomic
def registrar_venta(request):
    productos = Producto.objects.all()
    if request.method == 'POST':
        venta_form = SaleForm(request.POST)

        if venta_form.is_valid():
            venta = venta_form.save(commit=False)
            venta.total = 0
            venta.save()

            # Recorremos los productos del carrito manualmente
            producto_ids = request.POST.getlist('producto_id[]')
            cantidades = request.POST.getlist('cantidad[]')

            for pid, cant in zip(producto_ids, cantidades):
                try:
                    producto = Producto.objects.get(id=pid)
                except Producto.DoesNotExist:
                    continue

                cantidad = int(cant)

                # 🔹 Usar SIEMPRE el precio del inventario
                precio = producto.precio  
                subtotal = cantidad * precio

                # Crear detalle
                detalle = SaleItem.objects.create(
                    venta=venta,
                    producto=producto,
                    cantidad=cantidad,
                    precio=precio,  # Guardamos el precio de inventario
                    total=subtotal
                )

                # Descontar stock
                producto.cantidad -= cantidad
                producto.save()

                # Sumar al total de la venta
                venta.total += subtotal

            venta.save()
            messages.success(request, "Venta registrada correctamente.")
            return redirect('listar_ventas')
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        venta_form = SaleForm()

    return render(request, 'ventas/registrar_venta.html', {
        'venta_form': venta_form,
        'productos': productos
    })

@login_required
def listar_ventas(request):
    ventas = Sale.objects.all()
    return render(request, 'ventas/listar_ventas.html', {'ventas': ventas})
